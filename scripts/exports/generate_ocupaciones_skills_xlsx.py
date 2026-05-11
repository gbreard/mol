#!/usr/bin/env python3
"""
Genera Excel long-format con ocupaciones ESCO + skills asociadas.

Cruza:
  - database/embeddings/esco_occupations_full.json    (3046 ocupaciones base)
  - database/embeddings/esco_occupation_skills.json   (relaciones essential/optional)
  - database/embeddings/esco_skills_metadata_full.json (descripcion, type, broader)
  - exports/esco_skills_rdf_extract.json              (reuse_level, green)
  - exports/esco_categorias_L1_L2.xlsx                (descripciones L1, L2)
  - RDF directo                                       (description, altLabels, scopeNote,
                                                       regulatedProfessionNote, status)

Output: exports/ocupaciones_esco_con_skills.xlsx (~129K filas, long format)
"""

import xml.sax
import json
import time
from pathlib import Path

import openpyxl


PROJECT = Path("/mnt/d/OEDE/Webscrapping")
RDF_PATH = "/mnt/d/Trabajos en PY/EPH-ESCO/01_datos_originales/Tablas_esco/Data/esco-v1.2.0.rdf"

OCC_SKILLS_JSON   = PROJECT / "database/embeddings/esco_occupation_skills.json"
SKILLS_META_JSON  = PROJECT / "database/embeddings/esco_skills_metadata_full.json"
OCC_FULL_JSON     = PROJECT / "database/embeddings/esco_occupations_full.json"
SKILLS_RDF_EXTRA  = PROJECT / "exports/esco_skills_rdf_extract.json"
CATEGORIAS_XLSX   = PROJECT / "exports/esco_categorias_L1_L2.xlsx"

ENRICHED_CACHE    = PROJECT / "database/embeddings/esco_occupations_enriched.json"
OUTPUT_XLSX       = PROJECT / "exports/ocupaciones_esco_con_skills.xlsx"


REUSE_LABEL_ES = {
    "transversal": "Transversal",
    "cross-sector": "Inter-sectorial",
    "sector-specific": "Específica del sector",
    "occupation-specific": "Específica de la ocupación",
}

TYPE_LABEL_ES = {
    "skill": "Habilidad",
    "knowledge": "Conocimiento",
}

ISCO_1D_LABELS = {
    "0": "Ocupaciones militares",
    "1": "Directores y gerentes",
    "2": "Profesionales científicos e intelectuales",
    "3": "Técnicos y profesionales de nivel medio",
    "4": "Personal de apoyo administrativo",
    "5": "Trabajadores de los servicios y vendedores",
    "6": "Agricultores y trabajadores agropecuarios, forestales y pesqueros",
    "7": "Oficiales, operarios y artesanos",
    "8": "Operadores de instalaciones y máquinas y ensambladores",
    "9": "Ocupaciones elementales",
}


# =====================================================================
# RDF parser: campos enriquecidos por ocupación
# =====================================================================

class OccEnrichHandler(xml.sax.ContentHandler):
    """
    Stack-based handler. Soporta concepts anidados (broader inline):
    sólo captura campos del concept en el TOP del stack si es ocupación.
    """

    def __init__(self):
        super().__init__()
        # Stack: cada elemento es (uri, data_dict | None, is_occupation)
        self.concept_stack = []
        self.results = {}

        # Campos top-level del concept (description, scopeNote, altLabel, status)
        # solo se capturan cuando TOP del stack es la ocupación.
        self.field = None          # 'description' | 'scope' | 'altlabel' | 'status'
        self.text_buf = ""
        self.lang = None

        # Flag para regulatedProfessionNote: persiste a través de sub-conceptos
        # (porque está anidado dentro de AssociationObject)
        self.in_regulated_block = 0  # contador (puede haber varios anidados)

        self._tick = 0

    def _top(self):
        return self.concept_stack[-1] if self.concept_stack else (None, None, False)

    def _find_occ_data(self):
        """Busca la ocupación MÁS RECIENTE en el stack (la innermost).

        El RDF anida ocupaciones inline (skos:related, skos:broader, hasAssociation),
        y los campos como regulatedProfessionNote pertenecen a la ocupación más
        cercana en el stack, no a la raíz.
        """
        for _, data, is_occ in reversed(self.concept_stack):
            if is_occ:
                return data
        return None

    def startElement(self, name, attrs):
        if name == "skos:Concept":
            uri = attrs.get("rdf:about", "")
            is_occ = "/esco/occupation/" in uri
            data = None
            if is_occ:
                data = {
                    "uri": uri,
                    "description_es": "",
                    "scope_note_es": "",
                    "regulated_status": "",
                    "alt_labels_es": [],
                    "status": "",
                }
            else:
                # Capturar status regulated/unregulated/partially-regulated
                # cuando aparece dentro de un bloque regulatedProfessionNote
                if "/regulated-professions/" in uri and self.in_regulated_block > 0:
                    occ_data = self._find_occ_data()
                    if occ_data is not None and not occ_data["regulated_status"]:
                        occ_data["regulated_status"] = uri.rsplit("/", 1)[-1]
            self.concept_stack.append((uri, data, is_occ))
            return

        self.text_buf = ""

        # regulatedProfessionNote tiene dos formas en el RDF:
        # 1) Inline:    <esco:regulatedProfessionNote><skos:Concept rdf:about=".../X"/>...
        # 2) Resource:  <esco:regulatedProfessionNote rdf:resource=".../X"/>
        # La forma resource es la más común (una por ocupación).
        if name == "esco:regulatedProfessionNote":
            occ_data = self._find_occ_data()
            if occ_data is not None:
                # Forma 2: atributo rdf:resource
                resource = attrs.get("rdf:resource", "")
                if "/regulated-professions/" in resource and not occ_data["regulated_status"]:
                    occ_data["regulated_status"] = resource.rsplit("/", 1)[-1]
                # Forma 1: inline. Activamos el flag para capturar el sub-Concept.
                self.in_regulated_block += 1
            return

        # description / scopeNote / altLabel / status: solo cuando TOP es ocupación
        # (para no capturar de skills, AssociationObjects u otros sub-conceptos).
        _, _, top_is_occ = self._top()
        if not top_is_occ:
            return

        if name == "dct:description":
            self.field = "description"
            self.lang = None
        elif name == "skos:scopeNote":
            self.field = "scope"
            self.lang = None
        elif name == "skosXl:altLabel":
            self.field = "altlabel"
            self.lang = None
        elif name == "iso-thes:status":
            self.field = "status"
        elif name == "skosXl:literalForm":
            self.lang = attrs.get("xml:lang")

    def characters(self, content):
        self.text_buf += content

    def endElement(self, name):
        if name == "skos:Concept":
            uri, data, is_occ = self.concept_stack.pop()
            if is_occ:
                self.results[uri] = data
                self._tick += 1
                if self._tick % 500 == 0:
                    print(f"  {self._tick} ocupaciones enriquecidas...")
            return

        if name == "esco:regulatedProfessionNote":
            if self.in_regulated_block > 0:
                self.in_regulated_block -= 1
            self.text_buf = ""
            return

        _, data, top_is_occ = self._top()
        if not top_is_occ:
            self.text_buf = ""
            return

        if name == "iso-thes:status" and self.field == "status":
            data["status"] = self.text_buf.strip()
            self.field = None
            self.text_buf = ""
            return

        if name == "esco:language" and self.field in ("description", "scope"):
            self.lang = self.text_buf.strip()
            self.text_buf = ""
            return

        if name == "esco:nodeLiteral" and self.field in ("description", "scope"):
            if self.lang == "es":
                txt = self.text_buf.strip()
                if self.field == "description":
                    data["description_es"] = txt
                elif self.field == "scope":
                    data["scope_note_es"] = txt
            self.text_buf = ""
            return

        if name == "skosXl:literalForm" and self.field == "altlabel":
            if self.lang == "es":
                txt = self.text_buf.strip()
                if txt and txt not in data["alt_labels_es"]:
                    data["alt_labels_es"].append(txt)
            self.text_buf = ""
            return

        if name == "dct:description" and self.field == "description":
            self.field = None
        elif name == "skos:scopeNote" and self.field == "scope":
            self.field = None
        elif name == "skosXl:altLabel" and self.field == "altlabel":
            self.field = None

        self.text_buf = ""


def extract_enriched_from_rdf():
    if ENRICHED_CACHE.exists():
        print(f"Cache existe: {ENRICHED_CACHE.name}")
        with open(ENRICHED_CACHE) as f:
            return json.load(f)

    print(f"Parseando RDF: {RDF_PATH}")
    print(f"Tamaño: {Path(RDF_PATH).stat().st_size / 1024 / 1024:.0f} MB")
    print("Esto tarda 5-10 minutos...")

    handler = OccEnrichHandler()
    parser = xml.sax.make_parser()
    parser.setContentHandler(handler)

    start = time.time()
    parser.parse(RDF_PATH)
    elapsed = time.time() - start

    print(f"\nParseado en {elapsed:.0f}s")
    print(f"Ocupaciones enriquecidas: {len(handler.results)}")

    has_desc = sum(1 for d in handler.results.values() if d["description_es"])
    has_scope = sum(1 for d in handler.results.values() if d["scope_note_es"])
    reg_count = sum(1 for d in handler.results.values() if d["regulated_status"] == "regulated")
    unreg_count = sum(1 for d in handler.results.values() if d["regulated_status"] == "unregulated")
    partial_count = sum(1 for d in handler.results.values() if d["regulated_status"] == "partially-regulated")
    has_alt = sum(1 for d in handler.results.values() if d["alt_labels_es"])
    avg_alt = sum(len(d["alt_labels_es"]) for d in handler.results.values()) / max(len(handler.results), 1)
    print(f"  Con description (es): {has_desc}")
    print(f"  Con scope_note (es):  {has_scope}")
    print(f"  Reguladas:            {reg_count}")
    print(f"  Parcialmente regul.:  {partial_count}")
    print(f"  No reguladas:         {unreg_count}")
    print(f"  Con altLabels (es):   {has_alt}")
    print(f"  Promedio altLabels/ocup: {avg_alt:.1f}")

    ENRICHED_CACHE.parent.mkdir(parents=True, exist_ok=True)
    with open(ENRICHED_CACHE, "w", encoding="utf-8") as f:
        json.dump(handler.results, f, ensure_ascii=False, indent=2)
    print(f"\nGuardado: {ENRICHED_CACHE}")

    return handler.results


# =====================================================================
# Diccionarios L1/L2
# =====================================================================

def load_l1_l2_dict():
    wb = openpyxl.load_workbook(CATEGORIAS_XLSX, read_only=True)

    l1 = {}
    for row in list(wb["Categorías L1"].iter_rows(values_only=True))[1:]:
        if row[0]:
            l1[row[0]] = row[1] or ""

    l2 = {}
    for row in list(wb["Subcategorías L2"].iter_rows(values_only=True))[1:]:
        if row[2]:
            l2[row[2]] = row[3] or ""

    return l1, l2


# =====================================================================
# Main
# =====================================================================

def main():
    print("=" * 70)
    print("PASO 1: Datos enriquecidos del RDF")
    print("=" * 70)
    enriched = extract_enriched_from_rdf()

    print("\n" + "=" * 70)
    print("PASO 2: Cargar JSONs auxiliares")
    print("=" * 70)

    with open(OCC_SKILLS_JSON) as f:
        occ_skills = json.load(f)["occupation_skills"]
    print(f"Relaciones occupation→skills: {len(occ_skills)} ocupaciones")

    with open(SKILLS_META_JSON) as f:
        skills_list = json.load(f)
    skills_by_uri = {s["uri"]: s for s in skills_list}
    print(f"Skills metadata: {len(skills_by_uri)}")

    with open(OCC_FULL_JSON) as f:
        occ_full = json.load(f)
    occs = occ_full["occupations"]
    print(f"Ocupaciones base: {len(occs)}")

    with open(SKILLS_RDF_EXTRA) as f:
        rdf_extra = json.load(f)
    skills_reuse = rdf_extra["skills_reuse"]
    skills_green = set(rdf_extra["skills_green"])
    print(f"Skills con reuse_level: {len(skills_reuse)}")
    print(f"Skills green: {len(skills_green)}")

    l1_dict, l2_dict = load_l1_l2_dict()
    print(f"L1 entries: {len(l1_dict)} | L2 entries: {len(l2_dict)}")

    print("\n" + "=" * 70)
    print("PASO 3: Generar Excel")
    print("=" * 70)

    headers = [
        "ISCO 1d", "ISCO 1d label",
        "ISCO 4d", "ISCO 4d label",
        "Código ESCO", "Ocupación ESCO (es)",
        "Descripción ocupación", "Etiquetas alternativas (es)",
        "Scope note", "Profesión regulada", "Estado",
        "Tipo relación",
        "Skill label (es)", "Tipo skill",
        "L1 código", "L1 descripción",
        "L2 código", "L2 descripción",
        "Reuse level", "Green",
        "Descripción skill", "URI skill",
    ]

    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet("Ocupaciones x Skills")
    ws.append(headers)

    rows_written = 0
    occs_processed = 0
    occs_sin_skills = 0

    # Ordenar por código ESCO para que el Excel quede ordenado naturalmente
    sorted_occs = sorted(occs, key=lambda o: (o.get("isco_code", ""), o.get("esco_code", "")))

    for occ_base in sorted_occs:
        occs_processed += 1
        occ_uri = occ_base["uri"]

        skills_data = occ_skills.get(occ_uri)
        if not skills_data:
            occs_sin_skills += 1
            continue

        enriched_data = enriched.get(occ_uri, {})

        isco_4d = occ_base.get("isco_code", "")
        isco_4d_label = occ_base.get("isco_label", "")
        isco_1d = isco_4d[0] if isco_4d else ""
        isco_1d_label = ISCO_1D_LABELS.get(isco_1d, "")
        esco_code = occ_base.get("esco_code", "")
        esco_label = occ_base.get("esco_label", "")

        descripcion = enriched_data.get("description_es", "")
        alt_labels = "; ".join(enriched_data.get("alt_labels_es", []))
        scope_note = enriched_data.get("scope_note_es", "")
        regulated_raw = enriched_data.get("regulated_status", "")
        regulated = {
            "regulated": "Regulada",
            "unregulated": "No regulada",
            "partially-regulated": "Parcialmente regulada",
        }.get(regulated_raw, regulated_raw or "")
        status = enriched_data.get("status", "")

        for tipo_rel in ("essential", "optional"):
            tipo_rel_es = "Esencial" if tipo_rel == "essential" else "Optativa"
            for skill_rel in skills_data.get(tipo_rel, []):
                skill_uri = skill_rel.get("skill_uri", "")
                skill_label = skill_rel.get("skill_label", "")
                l1_code = skill_rel.get("L1", "")
                l2_code = skill_rel.get("L2", "")

                skill_meta = skills_by_uri.get(skill_uri, {})
                skill_type_raw = skill_meta.get("type", "")
                tipo_skill = TYPE_LABEL_ES.get(skill_type_raw, skill_type_raw)
                skill_desc = skill_meta.get("description", "")

                reuse_raw = skills_reuse.get(skill_uri, "")
                reuse_label = REUSE_LABEL_ES.get(reuse_raw, reuse_raw)
                green = "Sí" if skill_uri in skills_green else "No"

                ws.append([
                    isco_1d, isco_1d_label,
                    isco_4d, isco_4d_label,
                    esco_code, esco_label,
                    descripcion, alt_labels,
                    scope_note, regulated, status,
                    tipo_rel_es,
                    skill_label, tipo_skill,
                    l1_code, l1_dict.get(l1_code, ""),
                    l2_code, l2_dict.get(l2_code, ""),
                    reuse_label, green,
                    skill_desc, skill_uri,
                ])
                rows_written += 1

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_XLSX)

    print(f"\nFilas escritas:           {rows_written:>8,}")
    print(f"Ocupaciones procesadas:   {occs_processed:>8,}")
    print(f"Sin skills (saltadas):    {occs_sin_skills:>8,}")
    print(f"Ocupaciones con skills:   {occs_processed - occs_sin_skills:>8,}")
    size_mb = OUTPUT_XLSX.stat().st_size / 1024 / 1024
    print(f"\nGuardado: {OUTPUT_XLSX} ({size_mb:.1f} MB)")


if __name__ == "__main__":
    main()
