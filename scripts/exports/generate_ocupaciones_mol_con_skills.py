#!/usr/bin/env python3
"""
Genera exports/ocupaciones_mol_con_skills.xlsx — espejo del catálogo ESCO,
pero con lo que el pipeline MOL realmente extrajo.

3 hojas:
  1. "Ocupaciones MOL"      → 1 fila por ocupación ESCO con resumen MOL
  2. "Skills x Ocupacion"   → 1 fila por (ocupación matcheada, skill extraída)
  3. "Skills MOL globales"  → 1 fila por skill con métricas globales

Lee BD en modo read-only — no afecta pipeline en ejecución.
"""

import sqlite3
import json
import statistics
from pathlib import Path
from collections import defaultdict, Counter

import openpyxl

PROJECT = Path("/mnt/d/OEDE/Webscrapping")
DB_PATH = f"file:{PROJECT}/database/bumeran_scraping.db?mode=ro"

ESCO_OCC_FULL_JSON     = PROJECT / "database/embeddings/esco_occupations_full.json"
ESCO_OCC_SKILLS_JSON   = PROJECT / "database/embeddings/esco_occupation_skills.json"
ESCO_SKILLS_META_JSON  = PROJECT / "database/embeddings/esco_skills_metadata_full.json"
ESCO_RDF_EXTRA_JSON    = PROJECT / "exports/esco_skills_rdf_extract.json"
CATEGORIAS_L1L2_XLSX   = PROJECT / "exports/esco_categorias_L1_L2.xlsx"

OUTPUT_XLSX = PROJECT / "exports/ocupaciones_mol_con_skills.xlsx"

REUSE_LABEL_ES = {
    "transversal": "Transversal",
    "cross-sector": "Inter-sectorial",
    "sector-specific": "Específica del sector",
    "occupation-specific": "Específica de la ocupación",
}

ID_CAP = 200  # max IDs concatenados por celda (Excel limit ~32K chars)


def fmt_ids(ids_list, total_count):
    """Concatena IDs con cap. Si excede, agrega marcador '...+N más'."""
    if not ids_list:
        return ""
    head = ";".join(str(x) for x in ids_list[:ID_CAP])
    if total_count > ID_CAP:
        return f"{head};...+{total_count - ID_CAP} más"
    return head


def load_l1_l2_dict():
    wb = openpyxl.load_workbook(CATEGORIAS_L1L2_XLSX, read_only=True)
    l1 = {}
    for row in list(wb["Categorías L1"].iter_rows(values_only=True))[1:]:
        if row[0]:
            l1[row[0]] = row[1] or ""
    l2 = {}
    for row in list(wb["Subcategorías L2"].iter_rows(values_only=True))[1:]:
        if row[2]:
            l2[row[2]] = row[3] or ""
    return l1, l2


def main():
    # ===== Cargar catálogos ESCO =====
    print("Cargando catálogos ESCO...")
    with open(ESCO_OCC_FULL_JSON) as f:
        esco_occ = json.load(f)
    occs_by_uri = {o['uri']: o for o in esco_occ['occupations']}
    print(f"  ESCO ocupaciones: {len(occs_by_uri)}")

    with open(ESCO_SKILLS_META_JSON) as f:
        skills_meta = json.load(f)
    skill_by_uri = {s['uri']: s for s in skills_meta}
    print(f"  ESCO skills meta: {len(skill_by_uri)}")

    with open(ESCO_OCC_SKILLS_JSON) as f:
        occ_skills_data = json.load(f)
    occ_skills_map = occ_skills_data['occupation_skills']

    # Lookup (occ_uri, skill_uri) -> 'essential' | 'optional' | None
    occ_skill_relation = {}
    skills_in_any_catalog = set()
    isco4d_skills = defaultdict(set)
    for occ_uri, data in occ_skills_map.items():
        occ_meta = occs_by_uri.get(occ_uri, {})
        isco4 = (occ_meta.get('isco_code', '') or '')[:4]
        for s in data.get('essential', []):
            occ_skill_relation[(occ_uri, s['skill_uri'])] = 'essential'
            skills_in_any_catalog.add(s['skill_uri'])
            if isco4:
                isco4d_skills[isco4].add(s['skill_uri'])
        for s in data.get('optional', []):
            occ_skill_relation[(occ_uri, s['skill_uri'])] = 'optional'
            skills_in_any_catalog.add(s['skill_uri'])
            if isco4:
                isco4d_skills[isco4].add(s['skill_uri'])

    print(f"  Pares ocup-skill en catálogo: {len(occ_skill_relation):,}")
    print(f"  Skills en catálogo (cualquier ocup): {len(skills_in_any_catalog):,}")

    with open(ESCO_RDF_EXTRA_JSON) as f:
        rdf_extra = json.load(f)
    skills_reuse = rdf_extra['skills_reuse']
    skills_green = set(rdf_extra['skills_green'])

    l1_dict, l2_dict = load_l1_l2_dict()

    # ===== Conectar BD MOL =====
    print("\nConectando BD MOL (read-only)...")
    con = sqlite3.connect(DB_PATH, uri=True)
    cur = con.cursor()

    # ===== PASO 1: Mapeo oferta → ocupación matcheada =====
    print("Cargando ofertas matcheadas...")
    offer_to_occ = {}
    for row in cur.execute("""
        SELECT id_oferta, esco_occupation_uri
        FROM ofertas_esco_matching
        WHERE esco_occupation_uri IS NOT NULL
    """):
        offer_to_occ[row[0]] = row[1]
    print(f"  Ofertas con esco_uri: {len(offer_to_occ):,}")

    # Conteo de ofertas por ocupación + lista de IDs (capeada)
    n_offers_by_occ = Counter(offer_to_occ.values())
    offers_by_occ = defaultdict(list)
    for id_off, occ_uri in offer_to_occ.items():
        if len(offers_by_occ[occ_uri]) < ID_CAP:
            offers_by_occ[occ_uri].append(id_off)

    # ===== PASO 2: Cargar skills extraídas =====
    print("Cargando skills extraídas...")

    # occ_uri -> skill_uri -> {label, scores, count, ids (cap)}
    occ_skill_data = defaultdict(lambda: defaultdict(lambda: {
        'label': '',
        'scores': [],
        'count': 0,
        'ids': [],
    }))

    # Métricas globales por skill
    global_skill_data = defaultdict(lambda: {
        'label': '',
        'occs': set(),
        'count': 0,
        'scores': [],
        'ids': [],
    })

    n_processed = 0
    n_no_label = 0
    for row in cur.execute("""
        SELECT id_oferta, esco_skill_uri, esco_skill_label, match_score
        FROM ofertas_esco_skills_detalle
        WHERE esco_skill_uri IS NOT NULL AND esco_skill_uri != ''
    """):
        id_oferta, skill_uri, skill_label, score = row

        occ_uri = offer_to_occ.get(id_oferta)
        if not occ_uri:
            continue

        # Resolver label si está vacío (DIAG D)
        label = skill_label
        if not label:
            label = skill_by_uri.get(skill_uri, {}).get('label', '')
        if not label:
            n_no_label += 1
            continue

        d = occ_skill_data[occ_uri][skill_uri]
        d['label'] = label
        d['count'] += 1
        if score is not None:
            d['scores'].append(score)
        if len(d['ids']) < ID_CAP:
            d['ids'].append(id_oferta)

        gd = global_skill_data[skill_uri]
        gd['label'] = label
        gd['occs'].add(occ_uri)
        gd['count'] += 1
        if score is not None:
            gd['scores'].append(score)
        if len(gd['ids']) < ID_CAP:
            gd['ids'].append(id_oferta)

        n_processed += 1
        if n_processed % 200_000 == 0:
            print(f"  ... {n_processed:,} skills procesadas")

    print(f"  Total skills procesadas: {n_processed:,}")
    print(f"  Sin label resoluble:    {n_no_label:,}")
    print(f"  Ocupaciones con skills: {len(occ_skill_data):,}")
    print(f"  Skills únicas globales: {len(global_skill_data):,}")

    con.close()

    total_occs_with_offers = len([k for k, v in n_offers_by_occ.items() if v > 0])
    print(f"  Ocupaciones MOL con ≥1 oferta: {total_occs_with_offers:,}")

    # ===== PASO 3: Generar Excel =====
    print("\nGenerando Excel...")
    wb = openpyxl.Workbook(write_only=True)

    # ---------- HOJA 1: Ocupaciones MOL ----------
    ws1 = wb.create_sheet("Ocupaciones MOL")
    ws1.append([
        "ESCO URI", "ESCO code", "ESCO label", "ISCO 4d", "ISCO label",
        "N ofertas MOL", "N skills distintas",
        "% skills en catálogo ESCO de la ocupación",
        "Top-3 skills MOL (con freq)",
        "Top-3 huérfanas (no en cat. ESCO de la ocup)",
        f"id_ofertas (cap {ID_CAP})",
    ])

    sorted_occs = sorted(occs_by_uri.items(),
                         key=lambda x: ((x[1].get('isco_code') or ''),
                                         (x[1].get('esco_code') or '')))

    for occ_uri, occ_meta in sorted_occs:
        n_off = n_offers_by_occ.get(occ_uri, 0)
        skills_dict = occ_skill_data.get(occ_uri, {})
        n_skills = len(skills_dict)

        if n_skills > 0:
            in_catalog = sum(1 for su in skills_dict
                             if (occ_uri, su) in occ_skill_relation)
            pct_in = round(in_catalog / n_skills * 100, 1)
        else:
            pct_in = 0

        sorted_skills = sorted(skills_dict.items(), key=lambda x: -x[1]['count'])
        top3_str = "; ".join(f"{d['label']} ({d['count']})" for _, d in sorted_skills[:3])
        huerf = [(su, d) for su, d in sorted_skills
                 if (occ_uri, su) not in occ_skill_relation][:3]
        huerf_str = "; ".join(f"{d['label']} ({d['count']})" for _, d in huerf)

        ids_str = fmt_ids(offers_by_occ.get(occ_uri, []), n_off)

        ws1.append([
            occ_uri, occ_meta.get('esco_code', ''), occ_meta.get('esco_label', ''),
            occ_meta.get('isco_code', ''), occ_meta.get('isco_label', ''),
            n_off, n_skills, pct_in,
            top3_str, huerf_str,
            ids_str,
        ])

    print("  Hoja 1 OK")

    # ---------- HOJA 2: Skills x Ocupacion (long, TODO) ----------
    ws2 = wb.create_sheet("Skills x Ocupacion")
    ws2.append([
        "ESCO URI", "ESCO code", "ESCO label", "ISCO 4d",
        "N ofertas en la ocupación",
        "Skill label", "Skill URI",
        "N apariciones", "Cobertura %",
        "Score promedio", "Score mediano",
        "Posición ranking",
        "L1 código", "L1 descripción",
        "L2 código", "L2 descripción",
        "Reuse level", "Green",
        "Tipo relación con la ocup. matcheada",
        "Está en catálogo del ISCO 4d?",
        "% ocupaciones MOL con esta skill",
        f"id_ofertas (cap {ID_CAP})",
    ])

    rows_h2 = 0
    for occ_uri, skills_dict in occ_skill_data.items():
        occ_meta = occs_by_uri.get(occ_uri, {})
        n_off = n_offers_by_occ.get(occ_uri, 0)
        isco4 = (occ_meta.get('isco_code', '') or '')[:4]
        isco4_skills_set = isco4d_skills.get(isco4, set())

        sorted_skills = sorted(skills_dict.items(), key=lambda x: -x[1]['count'])
        for rank, (skill_uri, sd) in enumerate(sorted_skills, 1):
            sm = skill_by_uri.get(skill_uri, {})
            relation = occ_skill_relation.get((occ_uri, skill_uri))
            relation_es = {'essential': 'Esencial', 'optional': 'Optativa'}.get(relation, 'Huérfana')
            in_isco = "Sí" if skill_uri in isco4_skills_set else "No"

            n_occs_skill = len(global_skill_data[skill_uri]['occs'])
            magnetic_pct = round(n_occs_skill / total_occs_with_offers * 100, 2) \
                           if total_occs_with_offers else 0

            scores = sd['scores']
            score_avg = round(statistics.mean(scores), 3) if scores else 0
            score_med = round(statistics.median(scores), 3) if scores else 0

            l1 = sm.get('L1', '')
            l2 = sm.get('L2', '')
            reuse_raw = skills_reuse.get(skill_uri, '')
            reuse_label = REUSE_LABEL_ES.get(reuse_raw, reuse_raw)
            green = "Sí" if skill_uri in skills_green else "No"

            ws2.append([
                occ_uri, occ_meta.get('esco_code', ''), occ_meta.get('esco_label', ''),
                isco4, n_off,
                sd['label'], skill_uri,
                sd['count'],
                round(sd['count'] / n_off * 100, 2) if n_off else 0,
                score_avg, score_med, rank,
                l1, l1_dict.get(l1, ''),
                l2, l2_dict.get(l2, ''),
                reuse_label, green,
                relation_es, in_isco, magnetic_pct,
                fmt_ids(sd['ids'], sd['count']),
            ])
            rows_h2 += 1
        if rows_h2 % 100_000 == 0 and rows_h2 > 0:
            print(f"  Hoja 2: {rows_h2:,} filas escritas...")

    print(f"  Hoja 2 OK ({rows_h2:,} filas)")

    # ---------- HOJA 3: Skills MOL globales ----------
    ws3 = wb.create_sheet("Skills MOL globales")
    ws3.append([
        "Skill URI", "Skill label",
        "L1 código", "L1 descripción",
        "L2 código", "L2 descripción",
        "Reuse level", "Green",
        "N ocupaciones MOL distintas",
        "% ocupaciones MOL distintas",
        "Total apariciones",
        "Score promedio global",
        "Está en catálogo ESCO de alguna ocup?",
        f"id_ofertas (cap {ID_CAP})",
    ])

    for skill_uri, gd in sorted(global_skill_data.items(), key=lambda x: -len(x[1]['occs'])):
        sm = skill_by_uri.get(skill_uri, {})
        l1 = sm.get('L1', '')
        l2 = sm.get('L2', '')
        n_occs = len(gd['occs'])
        scores = gd['scores']
        score_avg = round(statistics.mean(scores), 3) if scores else 0

        ws3.append([
            skill_uri, gd['label'],
            l1, l1_dict.get(l1, ''),
            l2, l2_dict.get(l2, ''),
            REUSE_LABEL_ES.get(skills_reuse.get(skill_uri, ''), ''),
            "Sí" if skill_uri in skills_green else "No",
            n_occs,
            round(n_occs / total_occs_with_offers * 100, 2) if total_occs_with_offers else 0,
            gd['count'], score_avg,
            "Sí" if skill_uri in skills_in_any_catalog else "No",
            fmt_ids(gd['ids'], gd['count']),
        ])

    print("  Hoja 3 OK")

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_XLSX)
    size_mb = OUTPUT_XLSX.stat().st_size / 1024 / 1024
    print(f"\nGuardado: {OUTPUT_XLSX} ({size_mb:.1f} MB)")


if __name__ == "__main__":
    main()
