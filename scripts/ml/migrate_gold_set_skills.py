#!/usr/bin/env python3
"""
M-10 P2: Migrar skills del Gold Set desde Excel a Supabase.

Lee sheet 17_Skills_Completas_ESCO del Excel y matchea labels
con URIs de skills_embeddings por substring matching.

Uso:
    python scripts/ml/migrate_gold_set_skills.py
    python scripts/ml/migrate_gold_set_skills.py --dry-run
"""

import json
import argparse
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
EXCEL_PATH = PROJECT_ROOT / "docs" / "MOL_Gold_Set_49_Ofertas_Validacion (15-12).xlsx"


def load_skills_catalog(client):
    """Load all skill labels from Supabase for matching."""
    all_skills = {}  # lower_label → {uri, label}
    offset = 0
    while True:
        r = client.table('skills_embeddings').select('skill_uri,skill_label').range(offset, offset + 999).execute()
        for row in (r.data or []):
            key = row['skill_label'].lower().strip()
            all_skills[key] = {'uri': row['skill_uri'], 'label': row['skill_label']}
        if len(r.data or []) < 1000:
            break
        offset += 1000
    return all_skills


def find_uri_by_substring(label: str, catalog: dict) -> tuple:
    """Find URI by substring matching. Returns (uri, matched_label) or (None, None).

    If multiple matches, takes the shortest (most specific).
    """
    label_lower = label.lower().strip()

    # Exact match first
    if label_lower in catalog:
        return catalog[label_lower]['uri'], catalog[label_lower]['label']

    # Substring: excel label is contained in ESCO label, or vice versa
    candidates = []
    for sk_label, info in catalog.items():
        if label_lower in sk_label or sk_label in label_lower:
            candidates.append((sk_label, info))

    if not candidates:
        return None, None

    # Take shortest match (most specific)
    candidates.sort(key=lambda x: len(x[0]))
    return candidates[0][1]['uri'], candidates[0][1]['label']


def main():
    parser = argparse.ArgumentParser(description="M-10 P2: Migrate Gold Set skills")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    print("M-10 P2: Migrando skills del Gold Set")
    print("=" * 50)

    # Connect
    config = json.loads((PROJECT_ROOT / "config" / "supabase_config.json").read_text())
    from supabase import create_client
    client = create_client(config['url'], config['service_role_key'])

    # Load catalog
    print("Cargando catálogo de skills...")
    catalog = load_skills_catalog(client)
    print(f"  Skills en catálogo: {len(catalog)}")

    # Read Excel
    import openpyxl
    wb = openpyxl.load_workbook(str(EXCEL_PATH), read_only=True)
    ws = wb['17_Skills_Completas_ESCO']
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    rows = []
    for row in ws.iter_rows(min_row=2):
        vals = {h: c.value for h, c in zip(headers, row)}
        if vals.get('ESCO Skill') and vals.get('ID Oferta'):
            rows.append(vals)

    print(f"Skills en Excel: {len(rows)}")

    # Match URIs
    matched = 0
    unmatched = 0
    to_insert = []

    for vals in rows:
        label = vals['ESCO Skill']
        uri, matched_label = find_uri_by_substring(label, catalog)

        if uri:
            matched += 1
        else:
            unmatched += 1

        es_digital = str(vals.get('Es Digital', '')).lower() in ('sí', 'si', 'yes', 'true', '1')

        to_insert.append({
            'id_oferta': str(vals['ID Oferta']),
            'skill_label': label,
            'skill_uri': uri,
            'origen': vals.get('Origen'),
            'tipo_skill': vals.get('Tipo Skill'),
            'categoria': vals.get('Categoría') or vals.get('Categoria'),
            'es_digital': es_digital,
            'fuente': 'excel_dic2025',
        })

    print(f"\nMatcheo URIs:")
    print(f"  Con URI: {matched} ({matched/len(to_insert)*100:.1f}%)")
    print(f"  Sin URI: {unmatched} ({unmatched/len(to_insert)*100:.1f}%)")

    if args.dry_run:
        print("\n[DRY-RUN] No se insertan.")
        return

    # Insert in batches
    inserted = 0
    errors = 0
    batch_size = 50
    for i in range(0, len(to_insert), batch_size):
        batch = to_insert[i:i + batch_size]
        try:
            client.table('gold_set_skills').upsert(
                batch, on_conflict='id_oferta,skill_label'
            ).execute()
            inserted += len(batch)
        except Exception as e:
            # Try one by one
            for row in batch:
                try:
                    client.table('gold_set_skills').upsert(
                        row, on_conflict='id_oferta,skill_label'
                    ).execute()
                    inserted += 1
                except Exception as e2:
                    errors += 1
                    if errors <= 3:
                        print(f"  ERROR: {row['id_oferta']}/{row['skill_label'][:30]}: {e2}")

    print(f"\nResultado: {inserted} insertadas, {errors} errores")

    # Verify
    r = client.table('gold_set_skills').select('id', count='exact', head=True).execute()
    print(f"Total en gold_set_skills: {r.count}")


if __name__ == "__main__":
    main()
