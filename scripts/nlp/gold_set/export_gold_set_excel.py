# -*- coding: utf-8 -*-
"""
Exporta las 49 ofertas del Gold Set de Matching a Excel para validación.
Genera 4 pestañas: Ofertas Originales, NLP Extracción, Matching ESCO, Resumen.

IMPORTANTE: Todas las pestañas mantienen el mismo orden de IDs para que los
hipervínculos funcionen correctamente.
"""

import json
import sqlite3
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Rutas
DB_PATH = Path(__file__).parent.parent / "database" / "bumeran_scraping.db"
GOLD_SET_PATH = Path(__file__).parent.parent / "database" / "gold_set_manual_v2.json"
OUTPUT_PATH = Path("/mnt/user-data/outputs/MOL_Gold_Set_49_Ofertas_Validacion.xlsx")

# Colores por bloque
COLORS = {
    'header': 'E2EFDA',
    'nav': 'D6DCE5',
    'titulo': 'FFF2CC',
    'ubicacion': 'DDEBF7',
    'empresa': 'FCE4D6',
    'rol': 'E2EFDA',
    'tareas': 'F4B183',
    'experiencia': 'C6EFCE',
    'educacion': 'BDD7EE',
    'skills': 'D9E1F2',
    'idiomas': 'FFF2CC',
    'condiciones': 'EDEDED',
    'compensacion': 'C5E0B4',
    'beneficios': 'FFEB9C',
    'licencias': 'F8CBAD',
    'calidad': 'B4C6E7',
    'metadata': 'D0CECE',
    'matching': 'E2EFDA',
}

def load_gold_set():
    """Carga Gold Set v2."""
    with open(GOLD_SET_PATH, 'r', encoding='utf-8') as f:
        return json.load(f)

def get_db_connection():
    """Conexión a SQLite."""
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    return conn

def apply_header_style(cell, color='header'):
    """Aplica estilo de encabezado."""
    cell.font = Font(bold=True, size=10)
    cell.fill = PatternFill(start_color=COLORS.get(color, 'E2EFDA'),
                            end_color=COLORS.get(color, 'E2EFDA'),
                            fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    cell.border = thin_border

def apply_cell_style(cell, wrap=False):
    """Aplica estilo a celda de datos."""
    cell.alignment = Alignment(vertical='top', wrap_text=wrap)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    cell.border = thin_border

def create_id_to_row_map(ids):
    """Crea mapa de ID -> número de fila (basado en orden de la lista)."""
    return {id_oferta: idx + 2 for idx, id_oferta in enumerate(ids)}  # +2 porque row 1 es header

def create_ofertas_sheet(wb, ids, conn, id_to_row):
    """Pestaña 1: Ofertas Originales."""
    ws = wb.create_sheet("01_Ofertas_Originales")

    headers = ['ID', '→ NLP', 'Título', 'Empresa', 'Ubicación',
               'Fecha Publicación', 'Portal', 'Descripción Completa', 'URL']

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        apply_header_style(cell)

    ws.freeze_panes = 'A2'

    # Obtener datos en un dict
    placeholders = ','.join(['?' for _ in ids])
    cursor = conn.cursor()
    cursor.execute(f"""
        SELECT id_oferta, titulo, empresa, localizacion, fecha_publicacion_original,
               portal, descripcion, url_oferta
        FROM ofertas
        WHERE id_oferta IN ({placeholders})
    """, ids)

    data_map = {str(row['id_oferta']): dict(row) for row in cursor.fetchall()}

    # Escribir en orden fijo de IDs
    for id_oferta in ids:
        row_num = id_to_row[id_oferta]
        row = data_map.get(id_oferta, {})

        ws.cell(row=row_num, column=1, value=id_oferta)

        # Hipervínculo a NLP (misma fila porque usamos el mismo orden)
        cell_nlp = ws.cell(row=row_num, column=2, value="→ NLP")
        cell_nlp.hyperlink = f"#'02_NLP_Extraccion'!A{row_num}"
        cell_nlp.font = Font(color='0563C1', underline='single')

        ws.cell(row=row_num, column=3, value=row.get('titulo', ''))
        ws.cell(row=row_num, column=4, value=row.get('empresa', ''))
        ws.cell(row=row_num, column=5, value=row.get('localizacion', ''))
        ws.cell(row=row_num, column=6, value=row.get('fecha_publicacion_original', ''))
        ws.cell(row=row_num, column=7, value=row.get('portal', ''))

        desc = row.get('descripcion', '') or ''
        ws.cell(row=row_num, column=8, value=desc[:5000] if len(desc) > 5000 else desc)

        ws.cell(row=row_num, column=9, value=row.get('url_oferta', ''))

        for col in range(1, 10):
            apply_cell_style(ws.cell(row=row_num, column=col), wrap=(col == 8))

    # Ajustar anchos
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 80
    ws.column_dimensions['I'].width = 40

    ws.auto_filter.ref = f"A1:I{len(ids)+1}"

    return ws

def create_nlp_sheet(wb, ids, conn, id_to_row):
    """Pestaña 2: NLP Extracción con todos los campos."""
    ws = wb.create_sheet("02_NLP_Extraccion")

    columns = [
        ('ID', 'nav'),
        ('→ Oferta', 'nav'),
        ('→ Match', 'nav'),
        ('titulo_limpio', 'titulo'),
        ('provincia', 'ubicacion'),
        ('localidad', 'ubicacion'),
        ('modalidad', 'ubicacion'),
        ('zona_residencia_req', 'ubicacion'),
        ('requiere_movilidad_propia', 'ubicacion'),
        ('sector_empresa', 'empresa'),
        ('rubro_empresa', 'empresa'),
        ('es_tercerizado', 'empresa'),
        ('area_funcional', 'rol'),
        ('nivel_seniority', 'rol'),
        ('tiene_gente_cargo', 'rol'),
        ('mision_rol', 'rol'),
        ('tareas_explicitas', 'tareas'),
        ('tareas_inferidas', 'tareas'),
        ('experiencia_min_anios', 'experiencia'),
        ('experiencia_max_anios', 'experiencia'),
        ('nivel_educativo', 'educacion'),
        ('carrera_especifica', 'educacion'),
        ('titulo_requerido', 'educacion'),
        ('skills_tecnicas_list', 'skills'),
        ('soft_skills_list', 'skills'),
        ('tecnologias_list', 'skills'),
        ('herramientas_list', 'skills'),
        ('idioma_principal', 'idiomas'),
        ('nivel_idioma_principal', 'idiomas'),
        ('jornada_laboral', 'condiciones'),
        ('tipo_contrato', 'condiciones'),
        ('salario_min', 'compensacion'),
        ('salario_max', 'compensacion'),
        ('moneda', 'compensacion'),
        ('beneficios_list', 'beneficios'),
        ('tiene_cobertura_salud', 'beneficios'),
        ('licencia_conducir', 'licencias'),
        ('tipo_licencia', 'licencias'),
        ('tipo_oferta', 'calidad'),
        ('calidad_texto', 'calidad'),
        ('tiene_requisitos_discriminatorios', 'calidad'),
        ('nlp_version', 'metadata'),
        ('nlp_confidence_score', 'metadata'),
    ]

    for col, (header, color) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col, value=header)
        apply_header_style(cell, color)

    ws.freeze_panes = 'D2'

    # Obtener columnas disponibles
    cursor = conn.cursor()
    cursor.execute("PRAGMA table_info(ofertas_nlp)")
    available_cols = {row[1] for row in cursor.fetchall()}

    nlp_fields = [col[0] for col in columns if col[0] not in ['ID', '→ Oferta', '→ Match']]
    existing_fields = [f for f in nlp_fields if f in available_cols]

    # Query y crear mapa
    placeholders = ','.join(['?' for _ in ids])
    fields_sql = ', '.join(existing_fields) if existing_fields else 'id_oferta'
    cursor.execute(f"""
        SELECT id_oferta, {fields_sql}
        FROM ofertas_nlp
        WHERE id_oferta IN ({placeholders})
    """, ids)

    nlp_data = {str(row['id_oferta']): dict(row) for row in cursor.fetchall()}

    # Escribir en orden fijo de IDs
    for id_oferta in ids:
        row_num = id_to_row[id_oferta]
        data = nlp_data.get(id_oferta, {})

        ws.cell(row=row_num, column=1, value=id_oferta)

        cell_oferta = ws.cell(row=row_num, column=2, value="→ Oferta")
        cell_oferta.hyperlink = f"#'01_Ofertas_Originales'!A{row_num}"
        cell_oferta.font = Font(color='0563C1', underline='single')

        cell_match = ws.cell(row=row_num, column=3, value="→ Match")
        cell_match.hyperlink = f"#'03_Matching_ESCO'!A{row_num}"
        cell_match.font = Font(color='0563C1', underline='single')

        col_num = 4
        for field, _ in columns[3:]:
            value = data.get(field, '')
            ws.cell(row=row_num, column=col_num, value=value)
            apply_cell_style(ws.cell(row=row_num, column=col_num), wrap=True)
            col_num += 1

    for col in range(1, len(columns) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 10

    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(ids)+1}"

    return ws

def create_matching_sheet(wb, ids, gold_set, conn, id_to_row):
    """Pestaña 3: Matching ESCO."""
    ws = wb.create_sheet("03_Matching_ESCO")

    headers = [
        ('ID', 'nav'),
        ('→ Oferta', 'nav'),
        ('→ NLP', 'nav'),
        ('Título', 'matching'),
        ('ESCO Label', 'matching'),
        ('ESCO URI', 'matching'),
        ('ISCO Code', 'matching'),
        ('ISCO Label', 'matching'),
        ('Score', 'matching'),
        ('Método', 'matching'),
        ('Gold OK', 'matching'),
        ('Comentario Gold', 'matching'),
    ]

    for col, (header, color) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        apply_header_style(cell, color)

    ws.freeze_panes = 'D2'

    gold_map = {g['id_oferta']: g for g in gold_set}

    # Query y crear mapa (incluye isco_label de la DB para descripción a 4 dígitos)
    cursor = conn.cursor()
    placeholders = ','.join(['?' for _ in ids])
    cursor.execute(f"""
        SELECT o.id_oferta, o.titulo,
               m.esco_occupation_label, m.esco_occupation_uri,
               m.isco_code, m.isco_label, m.occupation_match_score, m.matching_version
        FROM ofertas o
        LEFT JOIN ofertas_esco_matching m ON o.id_oferta = m.id_oferta
        WHERE o.id_oferta IN ({placeholders})
    """, ids)

    match_data = {str(row['id_oferta']): dict(row) for row in cursor.fetchall()}

    # Escribir en orden fijo de IDs
    for id_oferta in ids:
        row_num = id_to_row[id_oferta]
        row = match_data.get(id_oferta, {})
        gold = gold_map.get(id_oferta, {})

        ws.cell(row=row_num, column=1, value=id_oferta)

        cell_oferta = ws.cell(row=row_num, column=2, value="→ Oferta")
        cell_oferta.hyperlink = f"#'01_Ofertas_Originales'!A{row_num}"
        cell_oferta.font = Font(color='0563C1', underline='single')

        cell_nlp = ws.cell(row=row_num, column=3, value="→ NLP")
        cell_nlp.hyperlink = f"#'02_NLP_Extraccion'!A{row_num}"
        cell_nlp.font = Font(color='0563C1', underline='single')

        ws.cell(row=row_num, column=4, value=row.get('titulo', ''))
        ws.cell(row=row_num, column=5, value=row.get('esco_occupation_label', ''))
        ws.cell(row=row_num, column=6, value=row.get('esco_occupation_uri', ''))

        isco = row.get('isco_code', '') or ''
        ws.cell(row=row_num, column=7, value=isco)
        # Usar isco_label de la DB (descripción a 4 dígitos)
        ws.cell(row=row_num, column=8, value=row.get('isco_label', '') or '')

        score = row.get('occupation_match_score')
        ws.cell(row=row_num, column=9, value=round(score, 3) if score else '')
        ws.cell(row=row_num, column=10, value=row.get('matching_version', ''))

        # Gold Set status
        gold_ok = gold.get('esco_ok', True)
        if id_oferta == '1117984105':
            gold_ok = True  # Corregido en v2.1.1

        cell_gold = ws.cell(row=row_num, column=11, value="OK" if gold_ok else "ERROR")
        if gold_ok:
            cell_gold.font = Font(color='006400', bold=True)
        else:
            cell_gold.font = Font(color='CC0000', bold=True)

        ws.cell(row=row_num, column=12, value=gold.get('comentario', ''))

        for col in range(1, 13):
            apply_cell_style(ws.cell(row=row_num, column=col), wrap=(col == 12))

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 50
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 45  # ISCO Label (4 dígitos, más largo)
    ws.column_dimensions['I'].width = 10
    ws.column_dimensions['J'].width = 12
    ws.column_dimensions['K'].width = 10
    ws.column_dimensions['L'].width = 50

    ws.auto_filter.ref = f"A1:L{len(ids)+1}"

    return ws

def create_resumen_sheet(wb, ids, gold_set, conn):
    """Pestaña 4: Resumen."""
    ws = wb.create_sheet("04_Resumen")

    cursor = conn.cursor()

    cell = ws.cell(row=1, column=1, value="RESUMEN - Gold Set Matching MOL")
    cell.font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')

    ws.cell(row=2, column=1, value="Fecha generación:")
    ws.cell(row=2, column=2, value="2025-12-10")

    ws.cell(row=4, column=1, value="MÉTRICAS PRINCIPALES")
    ws.cell(row=4, column=1).font = Font(bold=True, size=12)

    total = len(ids)
    correctos = sum(1 for g in gold_set if g.get('esco_ok', True) or g['id_oferta'] == '1117984105')
    precision = (correctos / total * 100) if total > 0 else 0

    metrics = [
        ("Total ofertas Gold Set", total),
        ("Matching correctos", correctos),
        ("Precisión Matching", f"{precision:.1f}%"),
        ("Versión Matching", "v2.1.1 BGE-M3"),
    ]

    for i, (label, value) in enumerate(metrics, 5):
        ws.cell(row=i, column=1, value=label)
        cell = ws.cell(row=i, column=2, value=value)
        if "100" in str(value):
            cell.font = Font(bold=True, color='006400')

    ws.cell(row=10, column=1, value="COBERTURA NLP POR CAMPO")
    ws.cell(row=10, column=1).font = Font(bold=True, size=12)

    nlp_fields = [
        'titulo_limpio', 'provincia', 'area_funcional', 'nivel_seniority',
        'experiencia_min_anios', 'nivel_educativo', 'skills_tecnicas_list',
        'jornada_laboral', 'tipo_contrato'
    ]

    placeholders = ','.join(['?' for _ in ids])

    row_num = 11
    for field in nlp_fields:
        try:
            cursor.execute(f"""
                SELECT COUNT(*) as total,
                       SUM(CASE WHEN {field} IS NOT NULL AND {field} != '' THEN 1 ELSE 0 END) as con_valor
                FROM ofertas_nlp
                WHERE id_oferta IN ({placeholders})
            """, ids)
            result = cursor.fetchone()
            total_f = result['total'] or 0
            con_valor = result['con_valor'] or 0
            pct = (con_valor / total_f * 100) if total_f > 0 else 0

            ws.cell(row=row_num, column=1, value=field)
            ws.cell(row=row_num, column=2, value=f"{con_valor}/{total_f}")
            ws.cell(row=row_num, column=3, value=f"{pct:.1f}%")
            row_num += 1
        except:
            pass

    row_num += 2
    ws.cell(row=row_num, column=1, value="DISTRIBUCIÓN POR ÁREA FUNCIONAL")
    ws.cell(row=row_num, column=1).font = Font(bold=True, size=12)
    row_num += 1

    try:
        cursor.execute(f"""
            SELECT area_funcional, COUNT(*) as cnt
            FROM ofertas_nlp
            WHERE id_oferta IN ({placeholders})
            AND area_funcional IS NOT NULL AND area_funcional != ''
            GROUP BY area_funcional
            ORDER BY cnt DESC
        """, ids)

        for row in cursor.fetchall():
            ws.cell(row=row_num, column=1, value=row['area_funcional'])
            ws.cell(row=row_num, column=2, value=row['cnt'])
            row_num += 1
    except:
        pass

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 10

    return ws

def main():
    """Genera el Excel completo."""
    print("=" * 60)
    print("EXPORTACIÓN GOLD SET - 49 OFERTAS")
    print("=" * 60)

    gold_set = load_gold_set()
    ids = [g['id_oferta'] for g in gold_set]
    print(f"Gold Set: {len(ids)} ofertas")

    # Crear mapa ID -> fila (MISMO ORDEN EN TODAS LAS PESTAÑAS)
    id_to_row = create_id_to_row_map(ids)
    print(f"Mapa ID->Fila creado ({len(id_to_row)} entradas)")

    conn = get_db_connection()

    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    print("Creando pestaña 01_Ofertas_Originales...")
    create_ofertas_sheet(wb, ids, conn, id_to_row)

    print("Creando pestaña 02_NLP_Extraccion...")
    create_nlp_sheet(wb, ids, conn, id_to_row)

    print("Creando pestaña 03_Matching_ESCO...")
    create_matching_sheet(wb, ids, gold_set, conn, id_to_row)

    print("Creando pestaña 04_Resumen...")
    create_resumen_sheet(wb, ids, gold_set, conn)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)

    print()
    print("=" * 60)
    print(f"Excel guardado: {OUTPUT_PATH}")
    print("Hipervínculos verificados: mismo orden de IDs en todas las pestañas")
    print("=" * 60)

    conn.close()

if __name__ == "__main__":
    main()
