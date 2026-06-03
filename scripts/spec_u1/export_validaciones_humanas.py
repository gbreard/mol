#!/usr/bin/env python3
"""
SPEC U-1 post-cierre — Exportar dataset de 218 validaciones humanas para SPEC W.

Output: data/spec_w/dataset_validaciones_humanas_2026_03_a_05.xlsx

Hojas:
  1. "Validaciones humanas" — 218 filas con datos enriquecidos
  2. "Resumen estadístico" — totales por validador, flag, timing, cola, conflicto
  3. "Casos conflicto temporal" — H18 (validación quedó huérfana post re-matching)
  4. "Patrón R240" — errores explícitos de Cyn apuntando a R240_operario_produccion
"""
import json, sqlite3
from datetime import datetime
from pathlib import Path
from collections import Counter
from supabase import create_client
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT = Path('/mnt/d/OEDE/Webscrapping')
OUT_DIR = ROOT / 'data/spec_w'
OUT_DIR.mkdir(parents=True, exist_ok=True)
OUT_PATH = OUT_DIR / 'dataset_validaciones_humanas_2026_03_a_05.xlsx'

SPEC_START = '2026-05-05'
SPEC_END = '2026-05-10'

VALIDADORES_HUMANOS = ['cinvazquez4@gmail.com', 'dschlese@trabajo.gob.ar', 'gbreard@gmail.com']


def fetch_supabase_validaciones(client):
    """Trae todas las validaciones humanas desde ofertas_dashboard."""
    rows = []
    page = 0
    while True:
        r = client.table('ofertas_dashboard').select(
            'id_oferta,titulo,descripcion,'
            'validacion_humana,validacion_humana_at,validacion_humana_por,validacion_correcciones,'
            'esco_occupation_uri,esco_occupation_label,isco_code,occupation_match_method'
        ).not_.is_('validacion_humana_at', 'null').in_('validacion_humana_por', VALIDADORES_HUMANOS).range(page*1000, (page+1)*1000-1).execute()
        if not r.data:
            break
        rows.extend(r.data)
        if len(r.data) < 1000:
            break
        page += 1
    return rows


def enrich_local(rows, db_path):
    """Enriquece cada fila con estado_validacion local, banderas, matching_timestamp."""
    con = sqlite3.connect(str(db_path))
    con.row_factory = sqlite3.Row

    ids = [r['id_oferta'] for r in rows]
    local_data = {}
    # batch in chunks of 500
    for i in range(0, len(ids), 500):
        chunk = ids[i:i+500]
        placeholders = ','.join('?' * len(chunk))
        for r in con.execute(f"""
            SELECT id_oferta, estado_validacion, notas_revision, bandera_spec_w_C1,
                   matching_version, matching_timestamp, esco_occupation_uri,
                   validado_timestamp, validado_por
            FROM ofertas_esco_matching WHERE id_oferta IN ({placeholders})
        """, chunk):
            local_data[r['id_oferta']] = dict(r)

    # validacion_historial: estado anterior por oferta (max-1 entry)
    historial = {}
    for i in range(0, len(ids), 500):
        chunk = ids[i:i+500]
        placeholders = ','.join('?' * len(chunk))
        # Tomar el primer estado_anterior por id (cronológicamente)
        for r in con.execute(f"""
            SELECT id_oferta, estado_anterior, MIN(timestamp) AS primer_ts
            FROM validacion_historial
            WHERE id_oferta IN ({placeholders})
            GROUP BY id_oferta
        """, chunk):
            historial[r['id_oferta']] = r['estado_anterior']

    con.close()
    return local_data, historial


def categorize_timing(vh_at_str):
    """pre_spec_u1 / durante_spec_u1 / post_spec_u1"""
    if not vh_at_str:
        return 'unknown'
    fecha = vh_at_str[:10]
    if fecha < SPEC_START:
        return 'pre_spec_u1'
    elif fecha <= SPEC_END:
        return 'durante_spec_u1'
    else:
        return 'post_spec_u1'


def determine_cola_origen(local, notas):
    """B2_subfaseD / C1_<tipo> / pre_spec / otro"""
    if notas and 'BANDERA_W: sub_ocupacion_bizarra_revisar' in notas:
        return 'B2_subfaseD'
    bandera_c1 = local.get('bandera_spec_w_C1') if local else None
    if bandera_c1:
        return f'C1_{bandera_c1}'
    estado = (local or {}).get('estado_validacion', '')
    if estado in ('pendiente_humano_subfaseD',):
        return 'subfaseD_sin_bandera'
    if estado in ('pendiente_humano_C1',):
        return 'C1_sin_bandera'
    if estado == 'validado_claude_C1':
        return 'auto_validada_C1'
    if estado == 'validado_claude_subfaseD':
        return 'auto_validada_subfaseD'
    if estado == 'validado_claude':
        return 'pre_spec_auto_validada'
    return 'otro'


def determine_estado_pre_spec(local_estado, historial_estado):
    """El estado que la oferta tenía pre-SPEC. Usamos primer entry en historial o estado actual."""
    # Si el estado actual es resultado del SPEC, buscar el original
    if local_estado in ('validado_claude_subfaseD', 'validado_claude_C1',
                         'pendiente_humano_subfaseD', 'pendiente_humano_C1'):
        return historial_estado or '(?)'
    return local_estado


def determine_re_matching(local, validacion_at_iso):
    """¿El matcher tocó la oferta DESPUÉS de la validación humana?
    Indicadores: estado_validacion es resultado SPEC, o matching_timestamp > validacion_at."""
    if not local:
        return False
    estado = local.get('estado_validacion', '')
    # Estados creados por SPEC = re-matching post-validación
    if estado in ('validado_claude_subfaseD', 'validado_claude_C1',
                   'pendiente_humano_subfaseD', 'pendiente_humano_C1'):
        return True
    # matching_timestamp posterior a validación
    mt = local.get('matching_timestamp')
    if mt and validacion_at_iso and mt[:19] > validacion_at_iso[:19]:
        return True
    return False


def determine_conflicto_temporal(rows_post, local, validacion_at_iso):
    """Conflicto temporal si:
    - Hubo re-matching post-validación, Y
    - La URI cambió (URI actual != URI que validó Cyn)
    Sin embargo, no tenemos la URI vista por Cyn explícitamente. Aproximación:
    si la nota de Cyn menciona un ISCO/label distinto al actual → conflicto.
    Más simple: si hubo re-matching POST validación, marcamos conflicto.
    """
    return determine_re_matching(local, validacion_at_iso)


def get_nota(validacion_correcciones):
    if not validacion_correcciones:
        return ''
    if isinstance(validacion_correcciones, dict):
        return validacion_correcciones.get('nota', '') or ''
    return str(validacion_correcciones)


def main():
    print("=== Bajando validaciones Supabase ===")
    config = json.loads((ROOT / 'config/supabase_config.json').read_text())
    client = create_client(config['url'], config['service_role_key'])
    rows = fetch_supabase_validaciones(client)
    print(f"  {len(rows)} validaciones de validadores humanos")

    print("\n=== Enriqueciendo con datos locales ===")
    local_data, historial = enrich_local(rows, ROOT / 'database/bumeran_scraping.db')
    print(f"  {len(local_data)} ofertas con datos locales")
    print(f"  {len(historial)} ofertas con validacion_historial")

    # Procesar cada fila
    enriched = []
    for r in rows:
        ls = local_data.get(r['id_oferta'], {})
        notas_local = ls.get('notas_revision') or ''
        vh_at = r.get('validacion_humana_at') or ''
        nota_cyn = get_nota(r.get('validacion_correcciones'))

        # Bandera SPEC W (subfaseD via notas)
        bandera_w = ''
        if 'BANDERA_W: sub_ocupacion_bizarra_revisar' in notas_local:
            bandera_w = 'sub_ocupacion_bizarra_revisar'

        bandera_c1 = ls.get('bandera_spec_w_C1') or ''

        re_matching = determine_re_matching(ls, vh_at)
        conflicto = determine_conflicto_temporal(rows, ls, vh_at)

        descripcion = (r.get('descripcion') or '')
        if len(descripcion) > 500:
            descripcion = descripcion[:500] + '...'

        enriched.append({
            'id_oferta': r['id_oferta'],
            'titulo': (r.get('titulo') or '')[:100],
            'descripcion_resumen': descripcion,
            'esco_uri_actual': r.get('esco_occupation_uri') or '',
            'esco_label_actual': r.get('esco_occupation_label') or '',
            'isco_code_actual': r.get('isco_code') or '',
            'estado_validacion_actual': ls.get('estado_validacion') or '',
            'bandera_spec_w': bandera_w,
            'bandera_spec_w_C1': bandera_c1,
            'validacion_humana_flag': r.get('validacion_humana') or '',
            'validacion_humana_at': vh_at,
            'validacion_humana_por': r.get('validacion_humana_por') or '',
            'nota_completa': nota_cyn,
            'timing_categoria': categorize_timing(vh_at),
            'cola_origen': determine_cola_origen(ls, notas_local),
            'estado_validacion_pre_spec': determine_estado_pre_spec(
                ls.get('estado_validacion', ''), historial.get(r['id_oferta'])
            ),
            'hubo_re_matching_post': 'sí' if re_matching else 'no',
            'conflicto_temporal': 'sí' if conflicto else 'no',
            'occupation_match_method': r.get('occupation_match_method') or '',
        })

    # Construir Excel
    print(f"\n=== Generando Excel: {OUT_PATH} ===")
    wb = Workbook()

    # === Hoja 1: Validaciones humanas ===
    ws = wb.active
    ws.title = 'Validaciones humanas'
    headers = [
        'id_oferta', 'titulo_oferta', 'descripcion_resumen',
        'esco_uri_actual', 'esco_label_actual', 'isco_code_actual',
        'estado_validacion_actual', 'bandera_spec_w', 'bandera_spec_w_C1',
        'validacion_humana_flag', 'validacion_humana_at', 'validacion_humana_por',
        'nota_completa',
        'timing_categoria', 'cola_origen', 'estado_validacion_pre_spec',
        'hubo_re_matching_post', 'conflicto_temporal',
    ]
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='305496', end_color='305496', fill_type='solid')
    border_thin = Border(left=Side(style='thin', color='CCCCCC'),
                          right=Side(style='thin', color='CCCCCC'),
                          top=Side(style='thin', color='CCCCCC'),
                          bottom=Side(style='thin', color='CCCCCC'))
    conflict_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    bandera_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center', wrap_text=True)
        c.border = border_thin

    field_order = [
        'id_oferta', 'titulo', 'descripcion_resumen',
        'esco_uri_actual', 'esco_label_actual', 'isco_code_actual',
        'estado_validacion_actual', 'bandera_spec_w', 'bandera_spec_w_C1',
        'validacion_humana_flag', 'validacion_humana_at', 'validacion_humana_por',
        'nota_completa',
        'timing_categoria', 'cola_origen', 'estado_validacion_pre_spec',
        'hubo_re_matching_post', 'conflicto_temporal',
    ]

    for row_idx, item in enumerate(enriched, 2):
        for col_idx, field in enumerate(field_order, 1):
            v = item[field]
            if isinstance(v, str) and len(v) > 32000:
                v = v[:32000] + '...[truncado]'
            c = ws.cell(row=row_idx, column=col_idx, value=v)
            c.alignment = Alignment(vertical='top', wrap_text=True)
            c.border = border_thin
            if field == 'conflicto_temporal' and v == 'sí':
                c.fill = conflict_fill
            if field in ('bandera_spec_w', 'bandera_spec_w_C1') and v:
                c.fill = bandera_fill

    # Anchos
    widths = {
        'A': 14, 'B': 35, 'C': 60, 'D': 40, 'E': 30, 'F': 8,
        'G': 22, 'H': 28, 'I': 32, 'J': 14, 'K': 18, 'L': 28,
        'M': 80, 'N': 16, 'O': 25, 'P': 22, 'Q': 14, 'R': 14,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = 'C2'

    # === Hoja 2: Resumen estadístico ===
    ws2 = wb.create_sheet('Resumen estadístico')
    ws2['A1'] = 'Resumen — dataset 218 validaciones humanas'
    ws2['A1'].font = Font(bold=True, size=14)
    row = 3

    def section(title):
        nonlocal row
        ws2.cell(row=row, column=1, value=title).font = Font(bold=True, size=12)
        row += 1

    def kv(k, v):
        nonlocal row
        ws2.cell(row=row, column=1, value=k)
        ws2.cell(row=row, column=2, value=v)
        row += 1

    section('Volumen')
    kv('Total validaciones', len(enriched))
    section('')
    section('Por validador')
    for k, n in Counter(e['validacion_humana_por'] for e in enriched).most_common():
        kv(k, n)
    section('')
    section('Por flag validacion_humana')
    for k, n in Counter(e['validacion_humana_flag'] for e in enriched).most_common():
        kv(k or '(vacío)', n)
    section('')
    section('Por timing_categoria')
    for k, n in Counter(e['timing_categoria'] for e in enriched).most_common():
        kv(k, n)
    section('')
    section('Por cola_origen')
    for k, n in Counter(e['cola_origen'] for e in enriched).most_common():
        kv(k, n)
    section('')
    section('Por estado_validacion_actual')
    for k, n in Counter(e['estado_validacion_actual'] for e in enriched).most_common():
        kv(k or '(vacío)', n)
    section('')
    n_conflict = sum(1 for e in enriched if e['conflicto_temporal'] == 'sí')
    n_bandera_w = sum(1 for e in enriched if e['bandera_spec_w'])
    n_bandera_c1 = sum(1 for e in enriched if e['bandera_spec_w_C1'])
    section('Banderas SPEC W')
    kv('Con bandera_spec_w (B2 subfaseD)', n_bandera_w)
    kv('Con bandera_spec_w_C1', n_bandera_c1)
    section('')
    section('Conflictos temporales (H18)')
    kv('Total con conflicto_temporal=sí', n_conflict)

    ws2.column_dimensions['A'].width = 40
    ws2.column_dimensions['B'].width = 30

    # === Hoja 3: Casos conflicto temporal ===
    ws3 = wb.create_sheet('Casos conflicto temporal')
    ws3['A1'] = 'Conflicto temporal H18 — validación quedó huérfana post re-matching'
    ws3['A1'].font = Font(bold=True, size=12)
    h3 = ['id_oferta', 'titulo', 'estado_actual', 'esco_label_actual',
          'fecha_validacion', 'validador', 'nota', 'estado_pre_spec', 'cola_origen']
    for ci, header in enumerate(h3, 1):
        c = ws3.cell(row=3, column=ci, value=header)
        c.font = header_font
        c.fill = header_fill
    row3 = 4
    for e in enriched:
        if e['conflicto_temporal'] != 'sí':
            continue
        ws3.cell(row=row3, column=1, value=e['id_oferta'])
        ws3.cell(row=row3, column=2, value=e['titulo'])
        ws3.cell(row=row3, column=3, value=e['estado_validacion_actual'])
        ws3.cell(row=row3, column=4, value=e['esco_label_actual'])
        ws3.cell(row=row3, column=5, value=e['validacion_humana_at'])
        ws3.cell(row=row3, column=6, value=e['validacion_humana_por'])
        nota = e['nota_completa'][:500] + ('...' if len(e['nota_completa']) > 500 else '')
        ws3.cell(row=row3, column=7, value=nota)
        ws3.cell(row=row3, column=8, value=e['estado_validacion_pre_spec'])
        ws3.cell(row=row3, column=9, value=e['cola_origen'])
        for col in range(1, 10):
            ws3.cell(row=row3, column=col).alignment = Alignment(vertical='top', wrap_text=True)
        row3 += 1
    for ci, w in enumerate([14, 35, 28, 30, 20, 28, 80, 22, 22], 1):
        ws3.column_dimensions[ws3.cell(row=3, column=ci).column_letter].width = w
    ws3.freeze_panes = 'A4'

    # === Hoja 4: Patrón R240 ===
    ws4 = wb.create_sheet('Patrón R240')
    ws4['A1'] = 'Errores explícitos apuntando a regla R240_operario_produccion'
    ws4['A1'].font = Font(bold=True, size=12)
    h4 = ['id_oferta', 'titulo', 'isco_actual', 'label_actual',
          'occupation_match_method', 'flag', 'nota']
    for ci, header in enumerate(h4, 1):
        c = ws4.cell(row=3, column=ci, value=header)
        c.font = header_font
        c.fill = header_fill
    row4 = 4
    for e in enriched:
        method = e.get('occupation_match_method', '') or ''
        is_r240 = 'R240' in method or 'operario_produccion' in method
        is_error = e['validacion_humana_flag'] in ('error',) or 'INCORRECTA' in e['nota_completa'].upper()
        if is_r240 or (is_error and '9329' in e['nota_completa']):
            ws4.cell(row=row4, column=1, value=e['id_oferta'])
            ws4.cell(row=row4, column=2, value=e['titulo'])
            ws4.cell(row=row4, column=3, value=e['isco_code_actual'])
            ws4.cell(row=row4, column=4, value=e['esco_label_actual'])
            ws4.cell(row=row4, column=5, value=method)
            ws4.cell(row=row4, column=6, value=e['validacion_humana_flag'])
            ws4.cell(row=row4, column=7, value=e['nota_completa'][:1000])
            for col in range(1, 8):
                ws4.cell(row=row4, column=col).alignment = Alignment(vertical='top', wrap_text=True)
            row4 += 1
    for ci, w in enumerate([14, 35, 10, 35, 35, 12, 80], 1):
        ws4.column_dimensions[ws4.cell(row=3, column=ci).column_letter].width = w
    ws4.freeze_panes = 'A4'

    wb.save(str(OUT_PATH))
    print(f"\n✅ Excel guardado: {OUT_PATH}")
    print(f"   Tamaño: {OUT_PATH.stat().st_size:,} bytes")
    print(f"\n=== Estadísticas finales ===")
    print(f"  Total filas: {len(enriched)}")
    print(f"  Con conflicto temporal: {n_conflict}")
    print(f"  Con bandera SPEC W (B2): {n_bandera_w}")
    print(f"  Con bandera_spec_w_C1: {n_bandera_c1}")
    print(f"  Por timing:")
    for k, n in Counter(e['timing_categoria'] for e in enriched).most_common():
        print(f"    {k}: {n}")


if __name__ == '__main__':
    main()
