#!/usr/bin/env python3
"""
SPEC U-1 sub-fase D Tarea adicional — Export Excel para validación humana de B2.

Genera Excel con 30 ofertas estratificadas del bucket B2 (BANDERA_W:
sub_ocupacion_bizarra_revisar) para que Cynthia/Diego revisen y aporten
correcciones.

Estrategia de muestreo:
  1. Cargar las 407 ofertas de B2 desde BD.
  2. Agrupar por ISCO 4-dig.
  3. Por cada ISCO: tomar min(5, n_ofertas).
  4. Si total > 30: recortar al azar manteniendo diversidad.
  5. Si total < 30: completar al azar de los pools restantes (no debería ocurrir
     con 407 ofertas y muchos ISCO distintos).

Hojas del Excel:
  - "Validacion B2"  — datos + columnas editables + dropdown evaluacion
  - "Instrucciones"  — texto para humano
  - "Metadata"       — totales + distribución + seed + bandera SQL

Output:
  - data/spec_u1/validacion_humana_B2_YYYYMMDD.xlsx
"""
import json
import random
import sqlite3
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path('/mnt/d/OEDE/Webscrapping')
DB_PATH = ROOT / 'database/bumeran_scraping.db'
OUT_DIR = ROOT / 'data/spec_u1'
SEED = 20260506

BANDERA_SQL = """SELECT om.id_oferta, om.isco_code, om.esco_occupation_label,
       om.esco_occupation_uri, om.occupation_match_method,
       o.titulo, o.descripcion, o.url_oferta, o.portal,
       n.titulo_limpio, n.tareas_explicitas, n.skills_tecnicas_list
FROM ofertas_esco_matching om
JOIN ofertas o ON o.id_oferta = om.id_oferta
LEFT JOIN ofertas_nlp n ON n.id_oferta = om.id_oferta
WHERE om.estado_validacion = 'pendiente_humano_subfaseD'
  AND om.notas_revision LIKE '%BANDERA_W: sub_ocupacion_bizarra_revisar%'
ORDER BY om.id_oferta"""


def fetch_b2(con):
    con.row_factory = sqlite3.Row
    rows = list(con.execute(BANDERA_SQL).fetchall())
    return rows


def estratificar_30(rows, seed=SEED):
    """Muestreo estratificado: máx 5 por ISCO, recortar al azar a 30."""
    rng = random.Random(seed)
    by_isco = defaultdict(list)
    for r in rows:
        by_isco[r['isco_code']].append(r)

    seleccion = []
    for isco, lista in by_isco.items():
        n_take = min(5, len(lista))
        seleccion.extend(rng.sample(lista, n_take))

    if len(seleccion) > 30:
        seleccion = rng.sample(seleccion, 30)
    elif len(seleccion) < 30:
        # completar de los pools que tienen >5
        sobrantes = []
        for isco, lista in by_isco.items():
            ya_tomados = [r for r in seleccion if r['isco_code'] == isco]
            sobrantes.extend([r for r in lista if r not in ya_tomados])
        rng.shuffle(sobrantes)
        seleccion.extend(sobrantes[:30 - len(seleccion)])

    rng.shuffle(seleccion)
    return seleccion


def truncate(s, n=500):
    if not s:
        return ''
    s = str(s)
    if len(s) <= n:
        return s
    return s[:n].rsplit(' ', 1)[0] + '...'


def parse_skills(skills_raw):
    """skills_tecnicas_list puede ser JSON list o string."""
    if not skills_raw:
        return ''
    try:
        skills = json.loads(skills_raw)
        if isinstance(skills, list):
            return ', '.join(str(s) for s in skills if s)
    except (json.JSONDecodeError, TypeError):
        pass
    return str(skills_raw)


def build_workbook(seleccion, total_b2, fecha_gen):
    wb = Workbook()

    # Hoja 1 — Validacion B2
    ws = wb.active
    ws.title = 'Validacion B2'

    headers = [
        'id_oferta', 'url_oferta', 'titulo', 'descripcion_resumen',
        'tareas_explicitas', 'skills_explicitas',
        'isco_asignado', 'label_asignado',
        'evaluacion', 'clasificacion_correcta_libre', 'comentario',
        'revisado_por', 'fecha_revision',
    ]

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='305496', end_color='305496', fill_type='solid')
    edit_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    border = Border(left=Side(style='thin', color='CCCCCC'),
                    right=Side(style='thin', color='CCCCCC'),
                    top=Side(style='thin', color='CCCCCC'),
                    bottom=Side(style='thin', color='CCCCCC'))

    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = border

    # Filas
    for row_idx, r in enumerate(seleccion, 2):
        titulo = r['titulo_limpio'] or r['titulo'] or ''
        desc = truncate(r['descripcion'], 500)
        tareas = r['tareas_explicitas'] or ''
        skills = parse_skills(r['skills_tecnicas_list'])

        valores = [
            r['id_oferta'],
            r['url_oferta'] or '',
            titulo,
            desc,
            tareas,
            skills,
            r['isco_code'] or '',
            r['esco_occupation_label'] or '',
            '',  # evaluacion
            '',  # clasificacion_correcta_libre
            '',  # comentario
            '',  # revisado_por
            '',  # fecha_revision
        ]
        for col_idx, v in enumerate(valores, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=v)
            c.alignment = Alignment(vertical='top', wrap_text=True)
            c.border = border
            if col_idx >= 9:  # columnas editables
                c.fill = edit_fill

    # Anchos
    widths = {
        'A': 14, 'B': 38, 'C': 36, 'D': 70, 'E': 50, 'F': 36,
        'G': 12, 'H': 32, 'I': 14, 'J': 38, 'K': 30, 'L': 16, 'M': 14,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.row_dimensions[1].height = 32

    # Dropdown evaluación
    dv = DataValidation(type='list', formula1='"OK,dudoso,mal"', allow_blank=True,
                         showDropDown=False, showErrorMessage=True,
                         errorTitle='Valor inválido', error='Usar OK / dudoso / mal')
    dv.add(f'I2:I{len(seleccion) + 1}')
    ws.add_data_validation(dv)

    # Freeze
    ws.freeze_panes = 'C2'

    # Hoja 2 — Instrucciones
    ws2 = wb.create_sheet('Instrucciones')
    instr = [
        ('Validación de clasificación — Bandera SPEC W', True, 16),
        ('', False, 11),
        ('Estas 30 ofertas fueron clasificadas por el sistema con una URI ESCO técnicamente correcta', False, 11),
        ('dentro del ISCO esperado, pero el label asignado es sospechoso de ser una sub-ocupación', False, 11),
        ('demasiado específica o errónea (ej: "comprador de café verde" para "administrativo de compras").', False, 11),
        ('', False, 11),
        ('Tu tarea:', True, 12),
        ('1. Lee título, descripción y tareas de la oferta (columnas C, D, E, F).', False, 11),
        ('2. En columna I (evaluacion): elegí del dropdown:', False, 11),
        ('     OK     — la clasificación asignada (G + H) tiene sentido para la oferta', False, 11),
        ('     dudoso — no estás seguro', False, 11),
        ('     mal    — claramente es errónea', False, 11),
        ('3. En columna J (clasificacion_correcta_libre): si marcaste "mal" o "dudoso", escribí qué', False, 11),
        ('   ocupación pensás que correspondería realmente (texto libre, no necesita formato ESCO).', False, 11),
        ('4. Columna K (comentario): cualquier observación adicional.', False, 11),
        ('5. Columnas L y M: tu nombre y fecha.', False, 11),
        ('', False, 11),
        ('No te preocupes por URIs ni códigos técnicos. Tu input es como humano que valida si la', False, 11),
        ('clasificación tiene sentido para el puesto laboral real.', False, 11),
        ('', False, 11),
        ('Cuando termines, devolveme el Excel completado.', True, 12),
    ]
    for idx, (text, bold, size) in enumerate(instr, 1):
        c = ws2.cell(row=idx, column=1, value=text)
        c.font = Font(bold=bold, size=size)
        c.alignment = Alignment(wrap_text=True)
    ws2.column_dimensions['A'].width = 100

    # Hoja 3 — Metadata
    ws3 = wb.create_sheet('Metadata')
    iscos_dist = defaultdict(int)
    for r in seleccion:
        iscos_dist[r['isco_code']] += 1

    rows3 = [
        ('Total ofertas en B2 (sub-ocupación bizarra)', total_b2),
        ('Muestra estratificada', len(seleccion)),
        ('Fecha de generación', fecha_gen),
        ('Seed determinístico', SEED),
        ('Bandera aplicada', "estado='pendiente_humano_subfaseD' AND notas_revision contiene 'BANDERA_W: sub_ocupacion_bizarra_revisar'"),
        ('SPEC origen', 'SPEC U-1 v3.1 sub-fase D (Tarea 4 Bucket 2)'),
        ('', ''),
        ('Distribución por ISCO 4-dig de la muestra:', ''),
    ]
    for isco, n in sorted(iscos_dist.items(), key=lambda x: -x[1]):
        rows3.append((f'  ISCO {isco}', n))

    for idx, (k, v) in enumerate(rows3, 1):
        c1 = ws3.cell(row=idx, column=1, value=k)
        c2 = ws3.cell(row=idx, column=2, value=v)
        if isinstance(v, int) or k == 'Bandera aplicada' or k == 'SPEC origen':
            pass
        if k.startswith('Distribución') or k.endswith(':'):
            c1.font = Font(bold=True)
        if k.startswith('  ISCO'):
            c1.font = Font(name='Consolas')
    ws3.column_dimensions['A'].width = 50
    ws3.column_dimensions['B'].width = 80

    return wb


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    fecha = datetime.now().strftime('%Y%m%d')
    fecha_human = datetime.now().strftime('%Y-%m-%d %H:%M')
    out_path = OUT_DIR / f'validacion_humana_B2_{fecha}.xlsx'

    print(f"Output: {out_path}")
    print(f"Seed: {SEED}")

    con = sqlite3.connect(str(DB_PATH))
    rows = fetch_b2(con)
    print(f"Total B2 en BD: {len(rows)}")
    if len(rows) == 0:
        raise RuntimeError("B2 vacío — verificar bandera SQL")

    seleccion = estratificar_30(rows, seed=SEED)
    print(f"Muestra estratificada: {len(seleccion)}")

    iscos = defaultdict(int)
    for r in seleccion:
        iscos[r['isco_code']] += 1
    print("Distribución muestra por ISCO:")
    for isco, n in sorted(iscos.items(), key=lambda x: -x[1]):
        print(f"  {isco}: {n}")

    wb = build_workbook(seleccion, total_b2=len(rows), fecha_gen=fecha_human)
    wb.save(str(out_path))
    print(f"\n✅ Excel guardado: {out_path}  ({out_path.stat().st_size} bytes)")
    con.close()
    return out_path, seleccion


if __name__ == '__main__':
    main()
